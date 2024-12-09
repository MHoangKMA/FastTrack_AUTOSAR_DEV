#!/usr/bin/env python3
# Shebang to specify the interpreter for running the script

import pandas as pd
import xml.etree.ElementTree as et
from pprint import pprint
import argparse
from tqdm import tqdm
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Importing necessary libraries:
# - pandas: for data manipulation and exporting to Excel.
# - xml.etree.ElementTree: for parsing the XML files.
# - pprint: for pretty-printing objects in the console.
# - argparse: for parsing command-line arguments.
# - tqdm: for displaying a progress bar.
# - os: for file and directory operations.
# - openpyxl: for manipulating Excel files and applying styles.

def main():
    # Main function to handle file processing and Excel report generation

    parser = argparse.ArgumentParser(description='Parse test results XML file to Excel', allow_abbrev=False)
    # Setting up the command-line argument parser. Description provides the script purpose, and allow_abbrev=False
    # ensures no abbreviations for arguments.

    parser.add_argument('input', type=str, metavar="convertXMLtoXLSX/ReportTest.xml", nargs="+", help="Location of the file(s) to process")
    # Adds positional argument 'input' to specify the XML file(s) to be processed.
    
    parser.add_argument('--output', "-o", type=str, metavar="report/ReportTest.xlsx", help="Location of the output file")
    # Adds optional argument '--output' to specify the location of the output Excel file.
    
    args = parser.parse_args()
    # Parse command-line arguments provided by thee user.

    outfile = args.output
    if not outfile:
        outfile = "{}".format(args.input[0].replace(".xml", ".xlsx"))
    # Determine the output file name. If not provided via --output, replace ".xml" in the input file name with ".xlsx".

    print(f"Output will go to: {outfile}")
    # Print the output file location for user information.

    # Ensure output directory exists
    output_dir = os.path.dirname(outfile)
    if not os.path.exists(output_dir) and output_dir != '':
        os.makedirs(output_dir)
    # Check if the output directory exists. If not, create it.

    # Define columns as per your data layout
    cols = [
        "tests", "failures", "disabled", "errors", "time", "timestamp",
        "name", "name2", "tests3", "failures4", "disabled5", "errors6", 
        "time7", "timestamp8", "name9", "status", "result", "time10", 
        "timestamp11", "classname", "failure", "message", "type"
    ]
    # Define the column headers for the resulting Excel file.

    rows = []
    # Initialize an empty list to store rows of data parsed from XML.

    totalcases = 0
    # Initialize a counter for the total number of test cases.

    # First pass to count total test cases for progress bar
    for f in args.input:
        if not os.path.isfile(f):
            print(f"File {f} does not exist. Skipping.")
            continue
        # Check if the input file exists. If not, skip to the next file.
        
        try:
            xtree = et.parse(f)
            xroot = xtree.getroot()
            testcases = xroot.findall('.//testcase')
            totalcases += len(testcases)
        # Parse the XML file, find all test cases, and count them.
        
        except Exception as e:
            print(f"Error parsing file {f}: {e}. Skipping.")
            continue
        # Handle any errors encountered during XML parsing.

    if totalcases == 0:
        print("No test cases found. Exiting.")
        return
    # If no test cases are found after processing all files, exit the program.

    bar = tqdm(total=totalcases, desc="Processing files")
    # Initialize a progress bar to track processing of test cases.

    # Second pass to process each test case
    for f in args.input:
        if not os.path.isfile(f):
            print(f"File {f} does not exist. Skipping.")
            continue
        # Check if the input file exists. If not, skip to the next file.

        try:
            xtree = et.parse(f)
            xroot = xtree.getroot()
            testcases = xroot.findall('.//testcase')
        # Parse the XML file, and extract test case elements.
        
        except Exception as e:
            print(f"Error parsing file {f}: {e}. Skipping.")
            continue
        # Handle errors in XML parsing.

        for testcase in testcases:
            testcase_name = testcase.attrib.get("name", "")
            status = testcase.attrib.get("status", "")
            result = testcase.attrib.get("result", "")
            time_val = testcase.attrib.get("time", "")
            timestamp = testcase.attrib.get("timestamp", "")
            classname = testcase.attrib.get("classname", "")
            # Extract attributes from each test case element (e.g., name, status, result).

            # Handle multiple failure elements
            failure_elems = testcase.findall("failure")
            failure_texts = []
            messages = []
            for failure in failure_elems:
                if failure.text:
                    failure_texts.append(failure.text)
                    messages.append(failure.text)  # Assuming 'message' is same as 'failure'
            # Extract failure information (if any) from each test case.
            # Aggregate failure texts into separate lists.

            failure_message = "\n\n".join(failure_texts)
            message = "\n\n".join(messages)
            # Combine multiple failure messages into a single string.

            # Populate row as per the new column layout
            row = {
                "tests": "", 
                "failures": "", 
                "disabled": "", 
                "errors": "", 
                "time": "", 
                "timestamp": "",
                "name": "", 
                "name2": testcase_name, 
                "tests3": "", 
                "failures4": "", 
                "disabled5": "", 
                "errors6": "", 
                "time7": time_val, 
                "timestamp8": timestamp, 
                "name9": classname,
                "status": status, 
                "result": result, 
                "time10": time_val, 
                "timestamp11": timestamp, 
                "classname": classname, 
                "failure": failure_message, 
                "message": message, 
                "type": ""
            }
            # Create a row dictionary containing the test case data.
            rows.append(row)
            # Append the row to the list of rows.
            
            bar.update(1)
            # Update the progress bar.

    bar.close()
    # Close the progress bar after all test cases are processed.

    print("\nWriting output...")
    # Notify the user that the output is being written.

    out_df = pd.DataFrame(rows, columns=cols)
    # Create a DataFrame from the list of rows and columns.

    try:
        out_df.to_excel(outfile, index=False)
    except Exception as e:
        print(f"Error writing to Excel file {outfile}: {e}")
        return
    # Write the DataFrame to an Excel file. If any errors occur, print an error message and exit.

    # Load the workbook and select the active worksheet
    wb = load_workbook(outfile)
    ws = wb.active
    # Open the generated Excel file and select the active sheet.

    # Apply styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    wrap_alignment = Alignment(wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    # Define styles for the header, text alignment, and cell borders.

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    # Apply the defined header styles to the first row (header row) in the worksheet.

    # Auto-fit columns based on the content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    # Automatically adjust the column widths based on the length of the data in each column.
    ws.column_dimensions['U'].width = 30  # Adjust as necessary
    ws.column_dimensions['V'].width = 30  # Adjust as necessary
    
    # Apply styles to all cells
    for row in ws.iter_rows(min_row=2, max_col=len(cols), max_row=ws.max_row):
        for cell in row:
            if cell.column_letter in ['U', 'V']:  # Assuming 'failure' is U and 'message' is V
                cell.alignment = wrap_alignment
            else:
                cell.alignment = center_alignment
            cell.border = thin_border
    # Apply text alignment and borders to all rows starting from the second row.

    # Freeze top row for better navigation
    ws.freeze_panes = "A2"
    # Freeze the top row in the worksheet to keep it visible during scrolling.

    # Save the styled workbook
    try:
        wb.save(outfile)
    except Exception as e:
        print(f"Error saving Excel file {outfile}: {e}")
        return
    # Save the Excel file. Handle any exceptions that occur during saving.

    pprint(out_df)
    # Pretty-print the DataFrame for debugging or user verification.

    print("Done")
    # Indicate completion of the process.

if __name__ == '__main__':
    main()
    # Entry point of the script. Execute the main() function when the script is run.
